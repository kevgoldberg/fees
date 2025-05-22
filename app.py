from flask import Flask, request, render_template, send_file, abort
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

app = Flask(__name__)

def process_workbook(stream, week: int) -> BytesIO:
    prev = max(1, week - 1)
    wb = load_workbook(stream)
    # Ensure table creation support
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    # Load worksheets
    cf = wb[f"Cash Funding Week {week}"]
    rb = wb[f"Report Batch Week {week}"]

    # Read data into DataFrames
    df_cf = pd.DataFrame(cf.values)
    df_cf.columns = df_cf.iloc[0]
    df_cf = df_cf[1:]
    df_rb = pd.DataFrame(rb.values)
    df_rb.columns = df_rb.iloc[0]
    df_rb = df_rb[1:]

    # Drop rows with Balance Due = 0
    df_cf = df_cf[df_cf["Balance Due"] != 0].copy()

    # Insert blank columns after Account Number
    idx = int(df_cf.columns.get_loc("Account Number")) + 1
    for col in ["Email Status","CK/SP/CO","Inv#/Item#/FileID#"]:
        df_cf.insert(idx, col, "")
        idx += 1
    # Sort by Annualized Fee % descending
    df_cf.sort_values(by="Annualized Fee %", ascending=False, inplace=True)

    # Populate Email Status from Report Batch
    df_cf["Email Status"] = df_cf["Account Number"].map(
        df_rb.set_index("Account Number")["Email Status"]
    )

    # Populate CK/SP/CO and Inv#/Item#/FileID# from prior week
    if week > 1:
        df_prev = pd.DataFrame(wb[f"Cash Funding Week {prev}"].values)
        df_prev.columns = df_prev.iloc[0]
        df_prev = df_prev[1:]
        for col in ["CK/SP/CO","Inv#/Item#/FileID#"]:
            df_cf[col] = df_cf["Account Number"].map(
                df_prev.set_index("Account Number")[col]
            )

    # Determine Qualified? from reference sheet
    ws_ref = wb["DO NOT DELETE"]
    quals = {row[2] for row in ws_ref.iter_rows(min_row=2, values_only=True) if row[2]}
    df_cf["Qualified?"] = df_cf["Account Type"].apply(
        lambda x: "Qualified" if x in quals else "Non-Qualified"
    )

    # Sort by Annualized Fee % descending
    df_cf.sort_values(by="Annualized Fee %", ascending=False, inplace=True)

    # Clear and rewrite Cash Funding sheet
    cf.delete_rows(1, cf.max_row)
    for r in dataframe_to_rows(df_cf, index=False, header=True):
        cf.append(r)
    # Create and style Excel table for Cash Funding
    cf_ref = f"A1:{get_column_letter(cf.max_column)}{cf.max_row}"
    tbl_cf = Table(displayName=f"CashFundingTable_Week{week}", ref=cf_ref)
    tbl_cf.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    cf.add_table(tbl_cf)
    # Bold header row and align left
    for cell in cf[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')

    # Rewrite Report Batch sheet as table
    rb.delete_rows(1, rb.max_row)
    for r in dataframe_to_rows(df_rb, index=False, header=True): rb.append(r)
    tbl_rb = Table(displayName=f"ReportBatchTable_Week{week}",
                   ref=f"A1:{get_column_letter(rb.max_column)}{rb.max_row}")
    tbl_rb.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    rb.add_table(tbl_rb)
    # Bold header row and align left on Report Batch
    for cell in rb[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')

    # Build header map for Cash Funding
    header_map = {cell.value: cell.column for cell in cf[1]}

    # Left-align all data in both tables
    align_left = Alignment(horizontal='left')
    for sheet in (cf, rb):
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = align_left

    # Currency format for columns I–L
    for col in ['I','J','K','L']:
        for cell in cf[col][1:]:
            cell.number_format = '$#,##0.00'

    # Conditional styling
    yellow = PatternFill('solid', fgColor='FFFF00')
    red_font = Font(color='FF0000')
    diff_col = header_map.get('Difference')
    pay_col = header_map.get('Pay Method')
    qual_col = header_map.get('Qualified?')
    fee_col = header_map.get('Annualized Fee %')
    for r in range(2, cf.max_row+1):
        fee_val = cf.cell(r, fee_col).value or 0
        if fee_val > 0.0267 and cf.cell(r, qual_col).value == 'Qualified':
            for c in range(1, cf.max_column+1): cf.cell(r,c).fill = yellow
        diff = cf.cell(r, diff_col).value
        if diff and diff < 0 and cf.cell(r, pay_col).value != 'Check':
            for c in range(1, cf.max_column+1): cf.cell(r,c).font = red_font

    # Manual column widths
    for name, width in [("Account Number",20), ("Household Full Name",42), ("Household Last Name",25)]:
        idx = header_map.get(name)
        cf.column_dimensions[get_column_letter(idx)].width = width

    # Hide reference sheet and freeze panes
    wb["DO NOT DELETE"].sheet_state = 'hidden'
    cf.freeze_panes = "A2"
    rb.freeze_panes = "A2"

    # Create pivot summaries using pandas
    df_all = df_cf.copy()
    # Check Fee Sweep: only 'Check' payments
    df_check = df_all[df_all['Pay Method'] == 'Check']
    pivot_check = df_check.pivot_table(index='Household Full Name', columns='CK/SP/CO', values='Balance Due', aggfunc='sum', fill_value=0)
    ws_check = wb.create_sheet('Check Fee Sweep')
    for r in dataframe_to_rows(pivot_check.reset_index(), index=False, header=True): ws_check.append(r)
    ws_check.freeze_panes = 'A2'
    ws_check.sheet_properties.tabColor = 'ADD8E6'
    # Client Fee Sweep: exclude 'Check'
    df_client = df_all[df_all['Pay Method'] != 'Check']
    pivot_client = df_client.pivot_table(index=['Household Full Name','Qualified?'], columns='CK/SP/CO', values='Balance Due', aggfunc='sum', fill_value=0)
    ws_client = wb.create_sheet('Client Fee Sweep')
    for r in dataframe_to_rows(pivot_client.reset_index(), index=False, header=True): ws_client.append(r)
    ws_client.freeze_panes = 'A3'
    ws_client.sheet_properties.tabColor = 'FDFD96'
    # Custodian Fee Sweep: filter fund families and exclude 'Check'
    df_cust = df_all[df_all['Fund Family'].isin(['Fidelity','Schwab']) & (df_all['Pay Method'] != 'Check')]
    pivot_cust = df_cust.pivot_table(index=['CK/SP/CO','Fund Family','Pay Method'], values='Balance Due', aggfunc='sum', fill_value=0)
    ws_cust = wb.create_sheet('Custodian Fee Sweep')
    for r in dataframe_to_rows(pivot_cust.reset_index(), index=False, header=True): ws_cust.append(r)
    ws_cust.freeze_panes = 'A4'
    ws_cust.sheet_properties.tabColor = '77DD77'
    # Save to BytesIO and return
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.route('/', methods=['GET', 'POST'])
def index():
    """Handle file upload and processing."""
    if request.method == 'POST':
        query_file = request.files.get('query_file')
        report_file = request.files.get('report_file')
        week = int(request.form.get('week', 0) or 0)

        # Validate files
        if not query_file or not query_file.filename:
            abort(400, 'Query file is required')
        if not report_file or not report_file.filename:
            abort(400, 'Report file is required')

        # Save the uploaded files into separate folders
        project_root = Path(__file__).resolve().parent
        qdir = project_root / 'Query'
        rdir = project_root / 'Report'
        qdir.mkdir(exist_ok=True)
        rdir.mkdir(exist_ok=True)
        qpath = qdir / query_file.filename
        query_file.save(str(qpath))
        rpath = rdir / report_file.filename
        report_file.save(str(rpath))

        # Process the query workbook
        with open(qpath, 'rb') as f:
            result = process_workbook(f, week)

        return send_file(
            result,
            as_attachment=True,
            download_name='processed.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
