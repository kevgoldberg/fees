from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def process_workbook(stream, week: int) -> BytesIO:
    prev = max(1, week - 1)
    wb = load_workbook(stream)

    cf = wb[f"Cash Funding Week {week}"]
    rb = wb[f"Report Batch Week {week}"]

    df_cf = pd.DataFrame(cf.values)
    df_cf.columns = df_cf.iloc[0]
    df_cf = df_cf[1:]
    df_rb = pd.DataFrame(rb.values)
    df_rb.columns = df_rb.iloc[0]
    df_rb = df_rb[1:]

    df_cf = df_cf[df_cf["Balance Due"] != 0].copy()

    idx = int(df_cf.columns.get_loc("Account Number")) + 1
    for col in ["Email Status", "CK/SP/CO", "Inv#/Item#/FileID#"]:
        df_cf.insert(idx, col, "")
        idx += 1
    df_cf.sort_values(by="Annualized Fee %", ascending=False, inplace=True)

    df_cf["Email Status"] = df_cf["Account Number"].map(
        df_rb.set_index("Account Number")["Email Status"]
    )

    if week > 1:
        df_prev = pd.DataFrame(wb[f"Cash Funding Week {prev}"].values)
        df_prev.columns = df_prev.iloc[0]
        df_prev = df_prev[1:]
        for col in ["CK/SP/CO", "Inv#/Item#/FileID#"]:
            df_cf[col] = df_cf["Account Number"].map(
                df_prev.set_index("Account Number")[col]
            )

    ws_ref = wb["DO NOT DELETE"]
    quals = {row[2] for row in ws_ref.iter_rows(min_row=2, values_only=True) if row[2]}
    df_cf["Qualified?"] = df_cf["Account Type"].apply(
        lambda x: "Qualified" if x in quals else "Non-Qualified"
    )

    df_cf.sort_values(by="Annualized Fee %", ascending=False, inplace=True)

    cf.delete_rows(1, cf.max_row)
    for r in dataframe_to_rows(df_cf, index=False, header=True):
        cf.append(r)
    cf_ref = f"A1:{get_column_letter(cf.max_column)}{cf.max_row}"
    tbl_cf = Table(displayName=f"CashFundingTable_Week{week}", ref=cf_ref)
    tbl_cf.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    cf.add_table(tbl_cf)

    for cell in cf[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')

    rb.delete_rows(1, rb.max_row)
    for r in dataframe_to_rows(df_rb, index=False, header=True):
        rb.append(r)
    tbl_rb = Table(displayName=f"ReportBatchTable_Week{week}",
                   ref=f"A1:{get_column_letter(rb.max_column)}{rb.max_row}")
    tbl_rb.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    rb.add_table(tbl_rb)

    for cell in rb[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')

    header_map = {cell.value: cell.column for cell in cf[1]}
    align_left = Alignment(horizontal='left')
    for sheet in (cf, rb):
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = align_left

    for col in ['I', 'J', 'K', 'L']:
        for cell in cf[col][1:]:
            cell.number_format = '$#,##0.00'

    yellow = PatternFill('solid', fgColor='FFFF00')
    red_font = Font(color='FF0000')
    diff_col = header_map.get('Difference')
    pay_col = header_map.get('Pay Method')
    qual_col = header_map.get('Qualified?')
    fee_col = header_map.get('Annualized Fee %')
    for r in range(2, cf.max_row + 1):
        fee_val = cf.cell(r, fee_col).value or 0
        if fee_val > 0.0267 and cf.cell(r, qual_col).value == 'Qualified':
            for c in range(1, cf.max_column + 1):
                cf.cell(r, c).fill = yellow
        diff = cf.cell(r, diff_col).value
        if diff and diff < 0 and cf.cell(r, pay_col).value != 'Check':
            for c in range(1, cf.max_column + 1):
                cf.cell(r, c).font = red_font

    for name, width in [("Account Number", 20), ("Household Full Name", 42), ("Household Last Name", 25)]:
        idx = header_map.get(name)
        cf.column_dimensions[get_column_letter(idx)].width = width

    wb["DO NOT DELETE"].sheet_state = 'hidden'
    cf.freeze_panes = "A2"
    rb.freeze_panes = "A2"

    df_all = df_cf.copy()
    df_check = df_all[df_all['Pay Method'] == 'Check']
    pivot_check = df_check.pivot_table(index='Household Full Name', columns='CK/SP/CO', values='Balance Due', aggfunc='sum', fill_value=0)
    ws_check = wb.create_sheet('Check Fee Sweep')
    for r in dataframe_to_rows(pivot_check.reset_index(), index=False, header=True):
        ws_check.append(r)
    ws_check.freeze_panes = 'A2'
    ws_check.sheet_properties.tabColor = 'ADD8E6'

    df_client = df_all[df_all['Pay Method'] != 'Check']
    pivot_client = df_client.pivot_table(index=['Household Full Name', 'Qualified?'], columns='CK/SP/CO', values='Balance Due', aggfunc='sum', fill_value=0)
    ws_client = wb.create_sheet('Client Fee Sweep')
    for r in dataframe_to_rows(pivot_client.reset_index(), index=False, header=True):
        ws_client.append(r)
    ws_client.freeze_panes = 'A3'
    ws_client.sheet_properties.tabColor = 'FDFD96'

    df_cust = df_all[df_all['Fund Family'].isin(['Fidelity', 'Schwab']) & (df_all['Pay Method'] != 'Check')]
    pivot_cust = df_cust.pivot_table(index=['CK/SP/CO', 'Fund Family', 'Pay Method'], values='Balance Due', aggfunc='sum', fill_value=0)
    ws_cust = wb.create_sheet('Custodian Fee Sweep')
    for r in dataframe_to_rows(pivot_cust.reset_index(), index=False, header=True):
        ws_cust.append(r)
    ws_cust.freeze_panes = 'A4'
    ws_cust.sheet_properties.tabColor = '77DD77'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output